# -*- coding: utf-8 -*-
# @__author__:choppa
# @DATA 2021/7/21
import json
from datetime import datetime

import openpyxl
import xlrd
import xlwt
from xlrd import xldate_as_tuple
from xlutils.copy import copy

from utils.dir_path import excel_dir
global str_obj

print('Excel数据处理')


# python 操作第三方文件--需要对应的库
# xlrd---.xls格式 使用openpyxl----.xlxs;大数据---pandas--.csv


def get_excel_data(sheetName, EXCElDIR, caseName=None):
    """
    :param EXCElDIR:
    :param sheetName: 表单名称
    :param caseName: 用例前缀
    :return:返回字典
    """
    relist = []
    # 1- excel 放到磁盘----把它加载到内存里
    workBook = xlrd.open_workbook(EXCElDIR, formatting_info=True)
    # 2- excel 文件对象--不止一个表--需要指定表
    workSheet = workBook.sheet_by_name(sheetName)
    ARGS = ['汽运任务', '分解任务', '火运计划']
    colindex = []
    for oneargs in ARGS:
        if oneargs in workSheet.row_values(0):
            colindex.append(workSheet.row_values(0).index(oneargs))
        else:
            print(f'【{oneargs}】表头不存在')

    idx = 0
    for _ in workSheet.col_values(0):
        if caseName is not None:
            getcoldata = []
            for num in colindex:
                temp = workSheet.cell(idx, num).value
                if is_json(temp):  # 需要转换成字典
                    temp = json.loads(temp)
                getcoldata.append(temp)
            relist.append(getcoldata)
        idx += 1
    return relist


def is_json(inStr):
    try:
        json.loads(inStr)
    except Exception as e:
        print(e)
        return False
    return True


def set_excel_data(excelDir):
    # 1- excel 放到磁盘----把它加载到内存里
    workBook = xlwt.Workbook(excelDir)
    workBookNew = copy(workBook)
    workSheetNew = workBookNew.get_sheet(0)
    return workBookNew, workSheetNew


def get_sheet1_data(excelDir, sheetName=None):
    # 1- excel 放到磁盘----把它加载到内存里
    workBook = xlrd.open_workbook(excelDir)
    # 2- excel 指定表
    sheet = workBook.sheet_by_name(sheetName)
    example_row = sheet.nrows
    data = []
    for eachrow in range(1, example_row):
        onerow = []
        for eachcol in range(3):  # 取3列的数据
            ctype = sheet.cell(eachrow, eachcol).ctype
            cell = sheet.cell_value(eachrow, eachcol)
            if ctype == 2 and cell % 1 == 0.0:  # ctype为2且为浮点
                cell = int(cell)  # 浮点转成整型
                cell = str(cell)  # 转成整型后再转成字符串，如果想要整型就去掉该行
            elif ctype == 3:
                date = datetime(*xldate_as_tuple(cell, 0))
                cell = date.strftime('%Y/%m/%d %H:%M:%S')
            elif ctype == 4:
                cell = True if cell == 1 else False
            onerow.append(cell)
        data.append(onerow)
    return data


def get_login_data(excelDir, sheetName="login"):
    # 1- excel 放到磁盘----把它加载到内存里
    global str_obj
    workBook = xlrd.open_workbook(excelDir)
    # 2- excel 指定表
    sheet = workBook.sheet_by_name(sheetName)
    nrows = sheet.nrows
    colnum = 3  # 列长度
    list1 = []
    num = 1
    for row_num in range(1, nrows):
        row_values = sheet.row_values(row_num)
        if row_values:
            str_obj = []
        # 获取每一行固定，从第一列1到多少列的数据，colnum 列数
        for i in range(colnum):
            ctype = sheet.cell(num, i).ctype
            cell = sheet.cell_value(num, i)
            if ctype == 2 and cell % 1 == 0.0:  # ctype为2且为浮点
                cell = int(cell)  # 浮点转成整型
                cell = str(cell)  # 转成整型后再转成字符串，如果想要整型就去掉该行
            elif ctype == 3:
                date = datetime(*xldate_as_tuple(cell, 0))
                cell = date.strftime('%Y/%m/%d %H:%M:%S')
            elif ctype == 4:
                cell = True if cell == 1 else False
            str_obj.append(cell)
        list1.append(str_obj)
        num = num + 1
    return list1


def get_ordercode_data(excelDir, sheetName="ordercode"):
    # 1- excel 放到磁盘----把它加载到内存里
    workBook = xlrd.open_workbook(excelDir)
    # 2- excel 文件对象--不止一个表--需要指定表
    sheet = workBook.sheet_by_name(sheetName)
    # 读取第1列的数据返回 第一行移除和空移除
    data = [sheet.col_values(0)[i] for i in range(1, len(sheet.col_values(0))) if sheet.col_values(0)[i] != ""]
    return data


def append_xlsx(path):
    """
    读取第3行的数据，不包括第3行最后一列的数据，循环写入到第4行至第23行
    :param path: 文件路径
    :return:
    """
    data = openpyxl.load_workbook(path)
    table = data.active
    # 1-获取第3行的数据
    val = get_sheet1_data(path)
    # 写入4-23行
    start_row = get_startrow(path)
    for item in range(start_row, start_row + 16):
        for j in range(len(val)):
            table.cell(item, j + 1).value = val[j]
    data.save(path)


def get_startrow(path, sheetName=""):
    """
    返回写入数据起始位置
    :param sheetName:
    :param path:
    :return:
    """
    workBook = xlrd.open_workbook(path)
    # 2- excel 指定表名
    sheet = workBook.sheet_by_name(sheetName)
    for i in range(sheet.nrows + 1):
        try:
            sheet.row_values(i)
        except Exception as e:
            print(e)
            return i


def get_row(path, sheetName='bind'):
    """
    返回写入数据起始位置
    :param sheetName:
    :param path:
    :return:
    """
    workBook = xlrd.open_workbook(path)
    # 2- excel 文件对象--不止一个表--需要指定表
    sheet = workBook.sheet_by_name(sheetName)
    for i in range(sheet.nrows + 1):
        try:
            sheet.row_values(i)
        except Exception as e:
            print(e)
            return i


def get_unbindrow(path, sheetName='unbind'):
    """
    返回写入数据起始位置
    :param sheetName: 表名
    :param path:
    :return:
    """
    workBook = xlrd.open_workbook(path)
    # 2- excel 文件对象--不止一个表--需要指定表
    sheet = workBook.sheet_by_name(sheetName)
    for i in range(sheet.nrows + 1):
        try:
            sheet.row_values(i)
        except Exception as e:
            print(e)
            return i


if __name__ == '__main__':
    ret = get_login_data(excel_dir, "usrdata")
    ret1 = get_sheet1_data(excel_dir, "usrdata")
    print(ret)
    print(ret1)
